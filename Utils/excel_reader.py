import xlrd2
from http_model.http_model import Model
from openpyxl import load_workbook


class ExcelReader:
    def Read(self):
        reader = xlrd2.open_workbook(r'.\data\case.xlsx')  # 创建excel对象
        sheet_names = reader.sheet_names()  # 获取所有sheet名
        models = []
        # 在sheet页中循环
        for name in sheet_names:
            #  定义sheet页变量并赋值
            sheet_name = reader.sheet_by_name(name)
            for i in range(sheet_name.nrows):
                if i == 0 or i == 1:  # 跳过前两行
                    continue
                data_list = []
                for j in range(sheet_name.ncols):
                    # if j == 0:  # 跳过第1列
                    #     continue
                    data_list.append(sheet_name.cell(i, j).value)
                model = Model()
                model.index = data_list[0]
                model.desc = data_list[1]
                model.url = data_list[2]
                model.method = data_list[3]
                model.headers = data_list[4]
                model.params = data_list[5]
                model.data = data_list[6]
                model.json = data_list[7]
                model.assert_data = data_list[8]
                model.assert_options = data_list[9]
                model.assert_value = data_list[10]
                model.extract = data_list[11]
                model.is_need = data_list[12]
                model.need_value = data_list[13]
                model.cell_index = data_list[15]
                model.feature = data_list[16]
                model.story = data_list[17]
                if type(model.index) is float:
                    models.append(model)
        return models

    def write_result(self, rows_num, result):
        workbook = load_workbook(filename="./data/case.xlsx")
        sheet = workbook.active
        sheet[f"O{rows_num}"] = result
        workbook.save(filename="./data/case.xlsx")

    def write_empty(self):
        workbook = load_workbook(filename="./data/case.xlsx")
        sheet = workbook.active
        i = 3
        while i < 1000:
            sheet[f"O{i}"] = ""
            i += 1
        workbook.save(filename="./data/case.xlsx")


if __name__ == '__main__':
    pass
